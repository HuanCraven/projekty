#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import json, re, glob
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------- load topics ----------
topics = []
for f in ["data/temata_sc4.json", "data/temata_sc1.json", "data/temata_sc5.json", "data/temata_sc8.json"]:
    topics += json.load(open(f, encoding="utf-8"))
topics.sort(key=lambda t: t["id"])
v3 = json.load(open("data/znalosti_v3.json", encoding="utf-8"))
def strip_src(z):
    z = re.sub(r"^ŠVP \(Ú[23/Ú]+\):\s*", "", z)
    return re.sub(r"^RVP – ", "", z)
print(f"Témat celkem: {len(topics)}")

# ---------- parse kameny from outlines ----------
kameny = {}  # code -> name
for f in sorted(glob.glob("podklady/SC?_outline.md")):
    txt = open(f, encoding="utf-8").read()
    for m in re.finditer(r"\*\*(\d+\.\d+\.\d+\.\d+)\.?\s+([^*—]+?)\*\*", txt):
        code, name = m.group(1), m.group(2).strip()
        kameny.setdefault(code, name)
print(f"Kamenů (balíčků) nalezeno v osnovách: {len(kameny)}")

# coverage from topics
cover_lead, cover_side = {}, {}
def code_of(s): return s.split(" ")[0].rstrip(".")
for t in topics:
    for k in t["vedouci"]:
        cover_lead.setdefault(code_of(k), []).append(t["id"])
    for k in t["vedlejsi"]:
        cover_side.setdefault(code_of(k), []).append(t["id"])

# sanity: codes referenced in topics but not found in outlines
refd = set(cover_lead) | set(cover_side)
unknown = sorted(refd - set(kameny))
if unknown: print("POZOR – kódy v tématech nenalezené v osnovách:", unknown)
uncovered = sorted(set(kameny) - refd)
print(f"Nepokryté kameny ({len(uncovered)}):", ", ".join(uncovered))

# ---------- workbook ----------
wb = Workbook()
ARIAL = "Arial"
hdr_fill = PatternFill("solid", start_color="1F4E5F")
grp_fills = {"SC4": "FDEBD0", "SC1": "D6EAF8", "SC5": "FDEDEC", "SC8": "D5F5E3"}
thin = Side(style="thin", color="B0B0B0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
GRP_NAZVY = {"SC4": "SC4 Rozvíjím svou odolnost", "SC1": "SC1 Umím se učit",
             "SC5": "SC5 Buduji dobré vztahy", "SC8": "SC8 Mám život ve svých rukou"}

ws = wb.active
ws.title = "Témata"
headers = ["Č.", "Téma", "Vedoucí ScioCíl", "Stěžejní kameny", "Vedlejší kameny",
           "Trojročí", "Anotace", "Náměty aktivit",
           "Tvrdé znalosti – příklady učiva",
           "Návaznost na projekty 2025/26", "Kdo si téma bere (vyplňte)"]
widths = [5, 26, 22, 34, 40, 10, 60, 70, 60, 24, 22]
for c, (h, w) in enumerate(zip(headers, widths), 1):
    cell = ws.cell(1, c, h)
    cell.font = Font(name=ARIAL, bold=True, color="FFFFFF", size=10)
    cell.fill = hdr_fill
    cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.column_dimensions[get_column_letter(c)].width = w
ws.freeze_panes = "C2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(topics)+1}"

def urov_text(u):
    """urovne je seznam trojročí, např. "1,2,3" nebo "3"."""
    t = sorted(x.strip() for x in str(u).split(",") if x.strip())
    if not t: return ""
    if len(t) == 1: return f"jen {t[0]}."
    if len(t) == 3: return "1.–3."
    return " i ".join(t)
for r, t in enumerate(topics, 2):
    vals = [t["id"], t["nazev"], GRP_NAZVY[t["skupina"]],
            "\n".join(t["vedouci"]), "\n".join(t["vedlejsi"]),
            urov_text(t["urovne"]), t["anotace"],
            "\n".join("• " + a for a in t["aktivity"]),
            "\n".join("• " + strip_src(z) for z in v3.get(str(t["id"]), {}).get("znalosti", [])),
            t["loni"] or "—", ""]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(r, c, v)
        cell.font = Font(name=ARIAL, size=9)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = border
        if c in (1, 6):
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="center")
    ws.cell(r, 3).fill = PatternFill("solid", start_color=grp_fills[t["skupina"]])

# ---------- sheet 2: pokrytí kamenů ----------
ws2 = wb.create_sheet("Pokrytí kamenů")
h2 = ["Kód kamene", "Název", "ScioCíl", "Stěžejní v tématech (č.)", "Vedlejší v tématech (č.)",
      "Počet stěžejní", "Počet vedlejší", "Celkem"]
w2 = [12, 46, 9, 26, 26, 13, 13, 10]
for c, (h, w) in enumerate(zip(h2, w2), 1):
    cell = ws2.cell(1, c, h)
    cell.font = Font(name=ARIAL, bold=True, color="FFFFFF", size=10)
    cell.fill = hdr_fill
    cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws2.column_dimensions[get_column_letter(c)].width = w
ws2.freeze_panes = "A2"

order = sorted(kameny, key=lambda k: [int(x) for x in k.split(".")])
r = 2
for code in order:
    lead = cover_lead.get(code, [])
    side = cover_side.get(code, [])
    sc = "SC" + code.split(".")[0]
    vals = [code, kameny[code], sc,
            ", ".join(map(str, lead)) or "—",
            ", ".join(map(str, side)) or "—",
            len(lead), len(side), len(lead) + len(side)]
    for c, v in enumerate(vals, 1):
        cell = ws2.cell(r, c, v)
        cell.font = Font(name=ARIAL, size=9)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = border
    if len(lead) + len(side) == 0:
        for c in range(1, 9):
            ws2.cell(r, c).fill = PatternFill("solid", start_color="F2F2F2")
    r += 1
ws2.auto_filter.ref = f"A1:H{r-1}"

# ---------- sheet 3: legenda ----------
ws3 = wb.create_sheet("Jak s tabulkou pracovat")
lines = [
    ("Katalog projektových témat – 1. pololetí 2026/27", True),
    ("", False),
    ("List „Témata“: 40 návrhů témat pro předmět Projekty, seskupených podle vedoucího ScioCíle (SC4, SC1, SC5, SC8).", False),
    ("Každé téma má stěžejní kameny (jádro projektu) a vedlejší kameny (přirozené přesahy, rozvíjené v menší míře).", False),
    ("Sloupec „Trojročí“: většina témat je společná pro 2. i 3. trojročí s diferenciací popsanou v katalogu (Word).", False),
    ("Sloupec „Kdo si téma bere“ je určen k vyplnění – zapište dvojici průvodců a termín bloku.", False),
    ("Sloupec „Tvrdé znalosti“: položky „ŠVP“ = učivo úrovní 2 a 3 z našeho ŠVP (jen SC1/SC4 – část ŠVP s cíli 5–8 nebyla v podkladu čitelná); položky „RVP“ = revidované RVP ZV (12/2024). Náměty, ne povinný výčet.", False),
    ("Filtrujte podle sloupce „Vedoucí ScioCíl“ nebo „Trojročí“ (řádek 1 má zapnutý filtr).", False),
    ("", False),
    ("List „Pokrytí kamenů“: přehled všech stavebních kamenů čtyř ScioCílů a témat, která je rozvíjejí.", False),
    ("Šedě podbarvené řádky = kameny zatím nepokryté žádným tématem (prostor pro další ročníky či volby průvodců).", False),
    ("Podrobné popisy témat včetně námětů aktivit a diferenciace najdete v dokumentu „Katalog projektových témat“ (Word).", False),
    ("Zdroj: upgrady ScioCílů 1, 4, 5, 8 (PREFINAL verze, 2025/26).", False),
]
for r, (txt, bold) in enumerate(lines, 1):
    cell = ws3.cell(r, 1, txt)
    cell.font = Font(name=ARIAL, size=12 if bold else 10, bold=bold)
ws3.column_dimensions["A"].width = 120

wb.save("vystupy/Projektova_temata.xlsx")
print("OK – uloženo vystupy/Projektova_temata.xlsx")
